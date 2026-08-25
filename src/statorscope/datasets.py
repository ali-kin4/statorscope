"""Loaders for public labelled induction-motor datasets.

Validating a fault detector against data you generated yourself proves only that
your generator and your detector share assumptions. These loaders pull independent,
published, labelled datasets so the claims in the README can be reproduced by
anyone.

Nothing here is downloaded automatically. Each loader takes a path to a file you
fetched yourself, and every dataset's citation and licence is recorded alongside
it -- see :data:`DATASETS`.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING

import numpy as np

from .signals import Motor, Recording

if TYPE_CHECKING:  # pragma: no cover
    from collections.abc import Iterator


@dataclass(frozen=True, slots=True)
class DatasetInfo:
    """Provenance for a public dataset. Cite it if you publish results."""

    key: str
    name: str
    url: str
    licence: str
    experimental: bool
    """False for simulation. Simulated data validates the maths, not the sensor."""

    citation: str


DATASETS: dict[str, DatasetInfo] = {
    "bbim2023": DatasetInfo(
        key="bbim2023",
        name="Broken Bars Induction Motors 2023 (BBIM2023)",
        url="https://data.mendeley.com/datasets/f5ksvkrp73/1",
        licence="CC BY 4.0",
        experimental=False,
        citation=(
            "Broken Bars Induction Motors 2023 (BBIM2023), Mendeley Data, V1, "
            "doi:10.17632/f5ksvkrp73.1 - simulated three-phase induction motors "
            "with 0-3 broken rotor bars at full load."
        ),
    ),
}

#: Motors described by the BBIM2023 workbooks, keyed by file stem.
BBIM2023_MOTORS: dict[str, Motor] = {
    "M24Ns_29Nb": Motor(pole_pairs=1, rotor_bars=29, line_hz=50.0),
    "M36Ns_44Nb": Motor(pole_pairs=2, rotor_bars=44, line_hz=60.0),
}

#: Worksheet name -> number of broken rotor bars.
BBIM2023_LABELS: dict[str, int] = {"H": 0, "1B": 1, "2B": 2, "3B": 3}

#: The simulations start from standstill; this much of the record is run-up.
BBIM2023_STARTUP_S = 0.6

BBIM2023_FS = 2000.0


@dataclass(frozen=True, slots=True)
class LabelledRecording:
    """A recording with its ground-truth label attached."""

    recording: Recording
    motor: Motor
    broken_bars: int
    true_slip: float | None
    """Slip computed from the simulator's rotor-speed channel, where present."""

    @property
    def healthy(self) -> bool:
        return self.broken_bars == 0


def load_bbim2023(
    path: str | Path,
    *,
    skip_startup_s: float = BBIM2023_STARTUP_S,
    sheets: tuple[str, ...] | None = None,
) -> Iterator[LabelledRecording]:
    """Load one BBIM2023 workbook, yielding one labelled recording per severity.

    Download the workbooks from https://data.mendeley.com/datasets/f5ksvkrp73/1
    (CC BY 4.0). Requires ``openpyxl``.

    The records begin at standstill, so ``skip_startup_s`` of run-up is discarded --
    steady-state MCSA assumes constant speed, and the acceleration would smear the
    carrier exactly the way a bad clock does.

    Args:
        path: Path to ``M24Ns_29Nb.xlsx`` or ``M36Ns_44Nb.xlsx``.
        skip_startup_s: Seconds of run-up to discard.
        sheets: Which severities to load; defaults to all of H, 1B, 2B, 3B.

    Yields:
        :class:`LabelledRecording`, one per severity level.

    Raises:
        ModuleNotFoundError: If ``openpyxl`` is not installed.
        KeyError: If the file stem is not a known BBIM2023 workbook.
    """
    try:
        import openpyxl  # noqa: PLC0415 - optional dependency, kept lazy on purpose
    except ModuleNotFoundError as exc:  # pragma: no cover - depends on env
        raise ModuleNotFoundError("load_bbim2023 needs openpyxl: pip install openpyxl") from exc

    p = Path(path)
    if p.stem not in BBIM2023_MOTORS:
        raise KeyError(
            f"{p.stem!r} is not a BBIM2023 workbook; expected one of {sorted(BBIM2023_MOTORS)}"
        )
    motor = BBIM2023_MOTORS[p.stem]
    skip = int(skip_startup_s * BBIM2023_FS)

    workbook = openpyxl.load_workbook(p, read_only=True)
    try:
        for sheet in sheets or tuple(BBIM2023_LABELS):
            if sheet not in workbook.sheetnames:
                continue
            rows = list(workbook[sheet].iter_rows(values_only=True))
            current = _columns(rows, (1, 2, 3))[skip:]
            # Only the 24-slot workbook carries a rotor-speed channel.
            speed = _columns(rows, (34,))[skip:, 0] if len(rows[0]) == 37 else None
            true_slip = None
            if speed is not None and np.isfinite(speed).any():
                synchronous = 2.0 * np.pi * motor.line_hz / motor.pole_pairs
                settled = float(np.nanmean(speed[-500:]))
                true_slip = (synchronous - settled) / synchronous
            yield LabelledRecording(
                recording=Recording.from_array(current, fs=BBIM2023_FS, name=f"{p.stem}/{sheet}"),
                motor=motor,
                broken_bars=BBIM2023_LABELS[sheet],
                true_slip=true_slip,
            )
    finally:
        workbook.close()


def _columns(rows: list[tuple[object, ...]], cols: tuple[int, ...]) -> np.ndarray:
    """Pull numeric columns out of a sheet that also contains text labels."""
    out = np.empty((len(rows), len(cols)), dtype=np.float64)
    for r, row in enumerate(rows):
        for c, col in enumerate(cols):
            try:
                out[r, c] = float(row[col])  # type: ignore[arg-type]
            except (TypeError, ValueError, IndexError):
                out[r, c] = np.nan
    return out
